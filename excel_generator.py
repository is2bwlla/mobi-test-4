import openpyxl

def create_excel_file(dados_adefa, dados_acara, dados_anfavea, dados_fenabrave):
    # Criando um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados de Mercado"
    
    # Cabeçalhos
    ws['A1'] = "Categoria"
    ws['B1'] = "Produção Ano Anterior"
    ws['C1'] = "Produção Ano Atual"
    ws['D1'] = "Exportação Ano Anterior"
    ws['E1'] = "Exportação Ano Atual"
    ws['F1'] = "Subtotal Produção"
    ws['G1'] = "Subtotal Exportação"
    
    # Verificando se existem dados de ADEFA e preenchendo
    if dados_adefa:
        row = 2
        # Produção
        if 'producao_ano_anterior' in dados_adefa and 'subtotal' in dados_adefa['producao_ano_anterior']:
            ws[f'B{row}'] = dados_adefa['producao_ano_anterior']['subtotal']
        if 'producao_ano_atual' in dados_adefa and 'subtotal' in dados_adefa['producao_ano_atual']:
            ws[f'C{row}'] = dados_adefa['producao_ano_atual']['subtotal']
        if 'exportacao_ano_anterior' in dados_adefa and 'subtotal' in dados_adefa['exportacao_ano_anterior']:
            ws[f'D{row}'] = dados_adefa['exportacao_ano_anterior']['subtotal']
        if 'exportacao_ano_atual' in dados_adefa and 'subtotal' in dados_adefa['exportacao_ano_atual']:
            ws[f'E{row}'] = dados_adefa['exportacao_ano_atual']['subtotal']
    
    # Verificando se existem dados de ACARA e preenchendo
    if dados_acara:
        row = 3
        if 'resumo_mercado' in dados_acara:
            ws[f'A{row}'] = "Resumo de Mercado"
            if 'acumulado_2025' in dados_acara['resumo_mercado']:
                ws[f'F{row}'] = dados_acara['resumo_mercado']['acumulado_2025']
            if 'acumulado_2024' in dados_acara['resumo_mercado']:
                ws[f'G{row}'] = dados_acara['resumo_mercado']['acumulado_2024']
    
    # Você pode adicionar verificações semelhantes para dados de ANFAVEA e FENABRAVE
    # Aqui estão exemplos para o ACARA, mas você pode adaptá-los para os outros prompts
    
    # Salvar o arquivo Excel
    wb.save("dados_mercado_gerado.xlsx")
    print("Arquivo Excel gerado com sucesso!")
